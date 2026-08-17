# generate_demo_workbook.py
# Builds a raw-vs-cleaned Excel workbook for presentation purposes only.
# Not part of the pipeline - re-run manually if cleaning logic changes.

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import openpyxl

RAW_FILE = "data/raw/KEIR8CFL.DTA"
OUT_FILE = "presentation/Raw_and_Cleaned_Full_Dataset.xlsx"

df = pd.read_stata(RAW_FILE, convert_categoricals=False)
meta = pd.io.stata.StataReader(RAW_FILE)
labels = meta.variable_labels()

# Restrict to adolescent girls
teens = df[(df["v012"] >= 15) & (df["v012"] <= 19)].copy()

raw_cols = ["caseid", "v012", "v024", "v025", "v106", "v149", "v190", "v501",
            "v525", "v384a", "v384b", "v384c", "v301", "v201", "v213", "v228",
            "v005", "v021", "v023"]

# Column definitions sheet, using labels pulled straight from the .dta file
defs = pd.DataFrame({
    "DHS Code": raw_cols,
    "Variable Label (from DHS Stata file)": [labels.get(c, "") for c in raw_cols],
})

# Raw sheet: untouched values, just readable headers
rename_map = {
    "caseid": "CaseID", "v012": "Age (v012)", "v024": "County_code (v024)",
    "v025": "Residence_code (v025)", "v106": "Education_code (v106)",
    "v149": "Education_detail_code (v149)", "v190": "Wealth_quintile (v190)",
    "v501": "Marital_status_code (v501)", "v525": "Age_at_first_sex_RAW (v525)",
    "v384a": "Heard_FP_radio_RAW (v384a)", "v384b": "Heard_FP_tv_RAW (v384b)",
    "v384c": "Heard_FP_print_RAW (v384c)", "v301": "Contraceptive_knowledge_code (v301)",
    "v201": "Children_ever_born (v201)", "v213": "Currently_pregnant (v213)",
    "v228": "Ever_terminated_pregnancy (v228)", "v005": "Survey_weight (v005)",
    "v021": "PSU (v021)", "v023": "Stratum (v023)",
}
raw_sheet = teens[raw_cols].rename(columns=rename_map)

# Cleaned sheet: same logic as build_analysis_dataset.py / feature_engineering.py
c = teens[raw_cols].copy()
c["ever_pregnant"] = ((c.v201.fillna(0) > 0) | (c.v213 == 1) | (c.v228 == 1)).astype(int)
c["ever_had_sex"] = (~c.v525.isin([0])).astype(int)
c["age_at_first_sex"] = c.v525.replace({0: np.nan, 49: np.nan})
c["urban"] = (c.v025 == 1).astype(int)
c["education"] = c.v106
c["wealth_quintile"] = c.v190
for col, label in [("v384a", "heard_fp_radio"), ("v384b", "heard_fp_tv"), ("v384c", "heard_fp_print")]:
    c[label] = c[col].fillna(-1).astype(int).map({0: "No", 1: "Yes", -1: "Not administered"})
c["knows_modern_contraception"] = (c.v301 == 3).astype(int)
marital_map = {0: "Never married", 1: "Married", 2: "Living together", 4: "Divorced/Separated", 5: "Widowed"}
c["marital_status"] = c.v501.map(marital_map)
c["county_code"] = c.v024

clean_cols = ["caseid", "county_code", "urban", "education", "wealth_quintile", "marital_status",
              "ever_had_sex", "age_at_first_sex", "heard_fp_radio", "heard_fp_tv",
              "heard_fp_print", "knows_modern_contraception", "ever_pregnant"]
clean_sheet = c[clean_cols].copy()
clean_sheet.columns = ["CaseID", "County_code", "Urban", "Education", "Wealth_quintile",
                        "Marital_status", "Ever_had_sex", "Age_at_first_sex",
                        "Heard_FP_radio", "Heard_FP_TV", "Heard_FP_print",
                        "Knows_modern_contraception", "Ever_pregnant (TARGET)"]

# Write workbook
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    defs.to_excel(writer, sheet_name="Column Definitions", index=False)
    raw_sheet.to_excel(writer, sheet_name="Raw Data - All Teens", index=False)
    clean_sheet.to_excel(writer, sheet_name="Cleaned Data - All Teens", index=False)

# Formatting: header style, frozen header row, column widths
wb = openpyxl.load_workbook(OUT_FILE)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        header_len = len(str(col_cells[0].value)) if col_cells[0].value else 10
        ws.column_dimensions[col_letter].width = min(max(header_len + 4, 14), 40)
wb.save(OUT_FILE)

print(f"Saved {len(raw_sheet):,} rows to {OUT_FILE}")